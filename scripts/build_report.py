#!/usr/bin/env python3
"""
MAVIS v7 — Report builder.

Reads the concordance CSV produced by apply_concordance_v4.py and writes a
multi-sheet XLSX with per-variant grades, threshold sensitivity analysis,
DDG distinguishability, mechanism consistency by class, and methodology
documentation.

USAGE
-----
    cd ~/mavis_v7
    python3 build_report.py

Reads:  results/mavis_v7_concordance.csv
Writes: results/mavis_v7_report.xlsx

Requires: openpyxl, pandas
"""
from __future__ import annotations
import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# Style definitions
# ============================================================================
FONT_HEADER  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_BODY    = Font(name="Arial", size=10)
FONT_BODY_B  = Font(name="Arial", size=10, bold=True)
FONT_NOTE    = Font(name="Arial", size=10, italic=True, color="555555")
FONT_TITLE   = Font(name="Arial", size=14, bold=True)

FILL_HEADER          = PatternFill("solid", start_color="2F4F4F")
FILL_SECTION         = PatternFill("solid", start_color="DDDDDD")
FILL_CONSISTENT      = PatternFill("solid", start_color="C6EFCE")
FILL_PARTIAL         = PatternFill("solid", start_color="FFEB9C")
FILL_INCONSISTENT    = PatternFill("solid", start_color="FFC7CE")
FILL_NA              = PatternFill("solid", start_color="EFEFEF")
FILL_FLAG_TRUE       = PatternFill("solid", start_color="D9EAD3")
FILL_FLAG_FALSE      = PatternFill("solid", start_color="F4CCCC")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header(cell):
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER


def style_body(cell, bold=False, align=ALIGN_CENTER):
    cell.font = FONT_BODY_B if bold else FONT_BODY
    cell.alignment = align
    cell.border = THIN_BORDER


def fill_for_grade(grade):
    return {
        "consistent":   FILL_CONSISTENT,
        "partial":      FILL_PARTIAL,
        "inconsistent": FILL_INCONSISTENT,
        "N/A":          FILL_NA,
    }.get(str(grade), None)


def fill_for_bool(val):
    return FILL_FLAG_TRUE if bool(val) else FILL_FLAG_FALSE


def write_header_row(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        style_header(c)
    ws.row_dimensions[row].height = 28


def auto_width(ws, max_width=40, min_width=10):
    """Set column widths based on max content length."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = min_width
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                v = str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# ============================================================================
# Sheet builders
# ============================================================================
def build_per_variant(wb, df, mode_A):
    ws = wb.create_sheet("Per_Variant")
    ws.cell(row=1, column=1, value="Per-Variant Detail").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    headers = [
        "system", "gene", "variant",
        "role", "phenotype",
        "expected_mech_class",
        "P1 mech t=1.0", "P1 mech t=1.5", "P1 mech t=2.0", "P1 mech t=2.5",
        "P1 grade t=1.0", "P1 grade t=1.5", "P1 grade t=2.0", "P1 grade t=2.5",
        "P1 summary (t=2.5)", "P1 stable",
        "P2 mech t=1.0", "P2 mech t=1.5", "P2 mech t=2.0", "P2 mech t=2.5",
        "P2 grade t=1.0", "P2 grade t=1.5", "P2 grade t=2.0", "P2 grade t=2.5",
        "P2 summary (t=2.5)", "P2 stable",
        "FP axes (t=1.0)", "MP axes (t=1.0)",
        "MAVIS tier (P1)", "Nbhd tier (P2)",
        "ddg_monomer", "ddg_monomer_sd",
        "max_abs_ddg",
        "concordance_strict_full", "concordance_relaxed_full",
        "structural_signal_strict", "structural_signal_relaxed",
        "external_consensus_strict", "external_consensus_relaxed",
        "AM class", "Franklin",
        "evidence_axes",
        "expected_ddg_monomer", "expected_ddg_fold_complex",
        "expected_ddg_binding", "expected_topology",
        "notes",
    ]
    if not mode_A:
        # Drop ground-truth columns in Mode B
        drop = {"role", "phenotype", "expected_mech_class",
                "P1 grade t=1.0", "P1 grade t=1.5", "P1 grade t=2.0", "P1 grade t=2.5",
                "P1 summary (t=2.5)", "P1 stable",
                "P2 grade t=1.0", "P2 grade t=1.5", "P2 grade t=2.0", "P2 grade t=2.5",
                "P2 summary (t=2.5)", "P2 stable",
                "FP axes (t=1.0)", "MP axes (t=1.0)",
                "evidence_axes",
                "expected_ddg_monomer", "expected_ddg_fold_complex",
                "expected_ddg_binding", "expected_topology"}
        headers = [h for h in headers if h not in drop]

    write_header_row(ws, 2, headers)

    # Map header → CSV column
    col_map = {
        "system": "system", "gene": "gene", "variant": "variant",
        "role": "role", "phenotype": "phenotype",
        "expected_mech_class": "expected_mech_class",
        "P1 mech t=1.0": "mech_t10", "P1 mech t=1.5": "mech_t15",
        "P1 mech t=2.0": "mech_t20", "P1 mech t=2.5": "mech_t25",
        "P1 grade t=1.0": "mech_consistency_t10", "P1 grade t=1.5": "mech_consistency_t15",
        "P1 grade t=2.0": "mech_consistency_t20", "P1 grade t=2.5": "mech_consistency_t25",
        "P1 summary (t=2.5)": "mech_consistency_summary",
        "P1 stable": "mech_consistency_threshold_stable",
        "P2 mech t=1.0": "nbhd_mech_t10", "P2 mech t=1.5": "nbhd_mech_t15",
        "P2 mech t=2.0": "nbhd_mech_t20", "P2 mech t=2.5": "nbhd_mech_t25",
        "P2 grade t=1.0": "nbhd_mech_consistency_t10",
        "P2 grade t=1.5": "nbhd_mech_consistency_t15",
        "P2 grade t=2.0": "nbhd_mech_consistency_t20",
        "P2 grade t=2.5": "nbhd_mech_consistency_t25",
        "P2 summary (t=2.5)": "nbhd_mech_consistency_summary",
        "P2 stable": "nbhd_mech_consistency_threshold_stable",
        "FP axes (t=1.0)": "mech_false_positive_axes_t10",
        "MP axes (t=1.0)": "mech_missed_positive_axes_t10",
        "MAVIS tier (P1)": "mavis_tier", "Nbhd tier (P2)": "nbhd_tier",
        "ddg_monomer": "ddg_monomer", "ddg_monomer_sd": "ddg_monomer_sd",
        "max_abs_ddg": "max_abs_ddg",
        "concordance_strict_full": "concordance_strict_full",
        "concordance_relaxed_full": "concordance_relaxed_full",
        "structural_signal_strict": "structural_signal_strict",
        "structural_signal_relaxed": "structural_signal_relaxed",
        "external_consensus_strict": "external_consensus_strict",
        "external_consensus_relaxed": "external_consensus_relaxed",
        "AM class": "AM class", "Franklin": "franklin",
        "evidence_axes": "evidence_axes",
        "expected_ddg_monomer": "expected_ddg_monomer",
        "expected_ddg_fold_complex": "expected_ddg_fold_complex",
        "expected_ddg_binding": "expected_ddg_binding",
        "expected_topology": "expected_topology",
        "notes": "notes",
    }

    grade_cols = {"P1 grade t=1.0", "P1 grade t=1.5", "P1 grade t=2.0",
                  "P1 grade t=2.5", "P1 summary (t=2.5)",
                  "P2 grade t=1.0", "P2 grade t=1.5", "P2 grade t=2.0",
                  "P2 grade t=2.5", "P2 summary (t=2.5)"}
    bool_cols  = {"P1 stable", "P2 stable"}

    row_start = 3
    for i, (_, r) in enumerate(df.iterrows()):
        for j, h in enumerate(headers, start=1):
            csv_col = col_map.get(h, h)
            val = r.get(csv_col, "")
            if pd.isna(val): val = ""
            cell = ws.cell(row=row_start + i, column=j, value=val)
            style_body(cell, align=ALIGN_LEFT if h in ("notes",) else ALIGN_CENTER)
            if h in grade_cols and val:
                fill = fill_for_grade(val)
                if fill: cell.fill = fill
            elif h in bool_cols and val != "":
                cell.fill = fill_for_bool(val)

    ws.freeze_panes = "D3"  # Freeze through "variant" column + first 2 rows
    auto_width(ws, max_width=35)
    # Notes column is wide
    notes_col = headers.index("notes") + 1
    ws.column_dimensions[get_column_letter(notes_col)].width = 60


def build_mech_consistency_summary(wb, df, mode_A):
    ws = wb.create_sheet("Mech_Consistency_Summary")
    ws.cell(row=1, column=1, value="Mechanism Consistency Summary").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    if not mode_A:
        ws.cell(row=3, column=1,
                value="Mode B: ground-truth annotations not present; mechanism "
                      "consistency cannot be computed.").font = FONT_NOTE
        return

    # Headline counts per pipeline per threshold
    ws.cell(row=3, column=1, value="Headline scores by pipeline × threshold").font = FONT_BODY_B
    headers = ["Pipeline", "Threshold", "Consistent", "Partial", "Inconsistent", "N/A", "Score"]
    write_header_row(ws, 4, headers)

    rows = []
    for pipeline_label, prefix in (("P1 (single-residue)", ""), ("P2 (neighborhood)", "nbhd_")):
        for thr_label, suf in (("t=1.0", "t10"), ("t=1.5", "t15"),
                               ("t=2.0", "t20"), ("t=2.5", "t25")):
            col = f"{prefix}mech_consistency_{suf}"
            counts = df[col].value_counts()
            n_c  = int(counts.get("consistent", 0))
            n_p  = int(counts.get("partial", 0))
            n_i  = int(counts.get("inconsistent", 0))
            n_na = int(counts.get("N/A", 0))
            n_graded = n_c + n_p + n_i
            score = (n_c + 0.5 * n_p) / n_graded if n_graded else 0.0
            rows.append((pipeline_label, thr_label, n_c, n_p, n_i, n_na, round(score, 3)))

    for i, vals in enumerate(rows, start=5):
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            style_body(cell, bold=(j == 7))

    # Per-class breakdown at t=2.5
    start_row = 5 + len(rows) + 2
    ws.cell(row=start_row, column=1,
            value="Per-class breakdown at t=2.5 (Sapozhnikov-confident threshold)").font = FONT_BODY_B
    write_header_row(ws, start_row + 1, ["Mech class", "n", "Consistent", "Partial",
                                          "Inconsistent", "N/A", "Score (P1)", "Score (P2)"])

    classes = sorted(df["expected_mech_class"].dropna().unique(),
                     key=lambda c: -len(df[df["expected_mech_class"] == c]))
    for k, cls in enumerate(classes):
        sub = df[df["expected_mech_class"] == cls]
        n = len(sub)
        c1 = sub["mech_consistency_t25"].value_counts()
        c2 = sub["nbhd_mech_consistency_t25"].value_counts()
        n_c = int(c1.get("consistent", 0))
        n_p = int(c1.get("partial", 0))
        n_i = int(c1.get("inconsistent", 0))
        n_na = int(c1.get("N/A", 0))
        n_g_p1 = n_c + n_p + n_i
        n_g_p2 = sum(int(c2.get(g, 0)) for g in ("consistent", "partial", "inconsistent"))
        score_p1 = (n_c + 0.5 * n_p) / n_g_p1 if n_g_p1 else 0.0
        score_p2 = ((int(c2.get("consistent", 0)) + 0.5 * int(c2.get("partial", 0)))
                    / n_g_p2 if n_g_p2 else 0.0)
        vals = (cls, n, n_c, n_p, n_i, n_na, round(score_p1, 3), round(score_p2, 3))
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=start_row + 2 + k, column=j, value=v)
            style_body(cell, bold=(j in (7, 8)))

    # Threshold-stable count
    stable_row = start_row + 2 + len(classes) + 2
    n_stable_p1 = int(df["mech_consistency_threshold_stable"].sum())
    n_stable_p2 = int(df["nbhd_mech_consistency_threshold_stable"].sum())
    ws.cell(row=stable_row, column=1,
            value="Threshold-stable variants (grade unchanged across all four thresholds)").font = FONT_BODY_B
    ws.cell(row=stable_row + 1, column=1, value=f"P1: {n_stable_p1}/{len(df)}")
    ws.cell(row=stable_row + 2, column=1, value=f"P2: {n_stable_p2}/{len(df)}")

    auto_width(ws, max_width=30)


def build_threshold_sensitivity(wb, df, mode_A):
    ws = wb.create_sheet("Threshold_Sensitivity")
    ws.cell(row=1, column=1, value="Threshold-Sensitive Variants").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    if not mode_A:
        ws.cell(row=3, column=1,
                value="Mode B: cannot identify threshold-sensitive variants without grades.").font = FONT_NOTE
        return

    note = ("Variants whose mechanism consistency grade changes across the four "
            "threshold values (t=1.0, 1.5, 2.0, 2.5 kcal/mol). These cases are "
            "where the threshold choice affects scientific interpretation.")
    ws.cell(row=3, column=1, value=note).font = FONT_NOTE
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)

    headers = ["gene", "variant", "expected_mech_class",
               "P1 t=1.0", "P1 t=1.5", "P1 t=2.0", "P1 t=2.5",
               "P2 t=1.0", "P2 t=1.5", "P2 t=2.0", "P2 t=2.5",
               "max_abs_ddg", "FP axes (t=1.0)", "MP axes (t=1.0)"]
    write_header_row(ws, 5, headers)

    unstable = df[~df["mech_consistency_threshold_stable"] | ~df["nbhd_mech_consistency_threshold_stable"]].copy()
    if len(unstable) == 0:
        ws.cell(row=6, column=1, value="(No threshold-sensitive variants — all 44 grades are stable)")
        return

    cols = [
        ("gene", "gene"), ("variant", "variant"),
        ("expected_mech_class", "expected_mech_class"),
        ("P1 t=1.0", "mech_consistency_t10"), ("P1 t=1.5", "mech_consistency_t15"),
        ("P1 t=2.0", "mech_consistency_t20"), ("P1 t=2.5", "mech_consistency_t25"),
        ("P2 t=1.0", "nbhd_mech_consistency_t10"), ("P2 t=1.5", "nbhd_mech_consistency_t15"),
        ("P2 t=2.0", "nbhd_mech_consistency_t20"), ("P2 t=2.5", "nbhd_mech_consistency_t25"),
        ("max_abs_ddg", "max_abs_ddg"),
        ("FP axes (t=1.0)", "mech_false_positive_axes_t10"),
        ("MP axes (t=1.0)", "mech_missed_positive_axes_t10"),
    ]
    grade_headers = {"P1 t=1.0", "P1 t=1.5", "P1 t=2.0", "P1 t=2.5",
                     "P2 t=1.0", "P2 t=1.5", "P2 t=2.0", "P2 t=2.5"}
    for i, (_, r) in enumerate(unstable.iterrows()):
        for j, (h, c) in enumerate(cols, start=1):
            v = r.get(c, "")
            if pd.isna(v): v = ""
            cell = ws.cell(row=6 + i, column=j, value=v)
            style_body(cell)
            if h in grade_headers and v:
                fill = fill_for_grade(v)
                if fill: cell.fill = fill
    ws.freeze_panes = "C6"
    auto_width(ws)


def build_distinguishability(wb, df):
    ws = wb.create_sheet("DDG_Distinguishability")
    ws.cell(row=1, column=1, value="DDG Distinguishability under Internal vs Sapozhnikov CIs").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    note = ("For each variant: at each threshold (0, 1.0, 1.5, 2.0, 2.5 kcal/mol), "
            "is the monomer ddg distinguishable from that magnitude? "
            "Internal CI uses run-to-run SD × 1.96 (tight). "
            "Sapozhnikov CI uses ±2.9 kcal/mol global bound (realistic).")
    ws.cell(row=3, column=1, value=note).font = FONT_NOTE
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)

    headers = ["gene", "variant", "ddg_monomer", "ddg_monomer_sd",
               "internal CI low", "internal CI high",
               "sapozhnikov CI low", "sapozhnikov CI high",
               "Int dist 0", "Int dist 1.0", "Int dist 1.5", "Int dist 2.0", "Int dist 2.5",
               "Sap dist 0", "Sap dist 1.0", "Sap dist 1.5", "Sap dist 2.0", "Sap dist 2.5"]
    write_header_row(ws, 5, headers)

    cols = [
        ("gene", "gene"), ("variant", "variant"),
        ("ddg_monomer", "ddg_monomer"), ("ddg_monomer_sd", "ddg_monomer_sd"),
        ("internal CI low", "ddg_monomer_ci95_internal_low"),
        ("internal CI high", "ddg_monomer_ci95_internal_high"),
        ("sapozhnikov CI low", "ddg_monomer_ci95_sapozhnikov_low"),
        ("sapozhnikov CI high", "ddg_monomer_ci95_sapozhnikov_high"),
        ("Int dist 0", "ddg_monomer_distinguishable_internal_from_0"),
        ("Int dist 1.0", "ddg_monomer_distinguishable_internal_from_t10"),
        ("Int dist 1.5", "ddg_monomer_distinguishable_internal_from_t15"),
        ("Int dist 2.0", "ddg_monomer_distinguishable_internal_from_t20"),
        ("Int dist 2.5", "ddg_monomer_distinguishable_internal_from_t25"),
        ("Sap dist 0", "ddg_monomer_distinguishable_sapozhnikov_from_0"),
        ("Sap dist 1.0", "ddg_monomer_distinguishable_sapozhnikov_from_t10"),
        ("Sap dist 1.5", "ddg_monomer_distinguishable_sapozhnikov_from_t15"),
        ("Sap dist 2.0", "ddg_monomer_distinguishable_sapozhnikov_from_t20"),
        ("Sap dist 2.5", "ddg_monomer_distinguishable_sapozhnikov_from_t25"),
    ]
    bool_headers = {h for h, _ in cols if "dist" in h}

    for i, (_, r) in enumerate(df.iterrows()):
        for j, (h, c) in enumerate(cols, start=1):
            v = r.get(c, "")
            if pd.isna(v): v = ""
            cell = ws.cell(row=6 + i, column=j, value=v)
            style_body(cell)
            if h in bool_headers and v != "":
                cell.fill = fill_for_bool(v)
    ws.freeze_panes = "C6"
    auto_width(ws)


def build_signal_in_benign(wb, df, mode_A):
    ws = wb.create_sheet("Signal_in_Benign")
    ws.cell(row=1, column=1,
            value="Structural Signal Detected in Benign Variants").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    if not mode_A:
        ws.cell(row=3, column=1,
                value="Mode B: cannot identify benign-vs-pathogenic split without role.").font = FONT_NOTE
        return

    note = ("Benign variants where MAVIS detects measurable structural signal "
            "(structural_signal_strict ≥ 1, or specific-axis mechanism call). "
            "These are NOT pipeline failures — they reflect that structural disruption "
            "and pathogenicity are distinct phenomena. Reported here for transparency.")
    ws.cell(row=3, column=1, value=note).font = FONT_NOTE
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)

    headers = ["gene", "variant", "role",
               "MAVIS mech (t=1.0)", "MAVIS mech (t=2.5)",
               "structural_signal_strict", "external_consensus_strict",
               "AM class", "Franklin", "notes"]
    write_header_row(ws, 5, headers)

    benign = df[df["role"] == "benign"].copy()
    # Identify benign variants firing structural signal
    def fires(row):
        sig = str(row.get("structural_signal_strict", "0/0"))
        try:
            n, d = sig.split("/")
            return int(n) >= 1
        except Exception:
            return False
    flagged = benign[benign.apply(fires, axis=1)].copy()
    if len(flagged) == 0:
        ws.cell(row=6, column=1,
                value="(No benign variants with structural signal detected)")
        return

    cols = [("gene", "gene"), ("variant", "variant"), ("role", "role"),
            ("MAVIS mech (t=1.0)", "mech_t10"), ("MAVIS mech (t=2.5)", "mech_t25"),
            ("structural_signal_strict", "structural_signal_strict"),
            ("external_consensus_strict", "external_consensus_strict"),
            ("AM class", "AM class"), ("Franklin", "franklin"),
            ("notes", "notes")]
    for i, (_, r) in enumerate(flagged.iterrows()):
        for j, (h, c) in enumerate(cols, start=1):
            v = r.get(c, "")
            if pd.isna(v): v = ""
            cell = ws.cell(row=6 + i, column=j,
                           value=v, )
            style_body(cell, align=ALIGN_LEFT if h == "notes" else ALIGN_CENTER)
    ws.freeze_panes = "C6"
    auto_width(ws, max_width=50)


def build_p1_vs_p2(wb, df, mode_A):
    ws = wb.create_sheet("P1_vs_P2_Comparison")
    ws.cell(row=1, column=1,
            value="Pipeline 1 vs Pipeline 2 Mechanism Comparison").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    note = ("Variants where P1 (single-residue) and P2 (neighborhood-windowed) "
            "produce different mechanism calls at t=1.0. These are cases where "
            "regional structural context changes the call.")
    ws.cell(row=3, column=1, value=note).font = FONT_NOTE
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)

    headers = ["gene", "variant", "expected_mech_class",
               "P1 mech (t=1.0)", "P2 mech (t=1.0)",
               "P1 grade (t=1.0)", "P2 grade (t=1.0)",
               "P1 tier", "P2 tier", "pipeline_agreement"]
    write_header_row(ws, 5, headers)

    diff = df[df["mech_t10"] != df["nbhd_mech_t10"]].copy()
    cols = [("gene", "gene"), ("variant", "variant"),
            ("expected_mech_class", "expected_mech_class"),
            ("P1 mech (t=1.0)", "mech_t10"), ("P2 mech (t=1.0)", "nbhd_mech_t10"),
            ("P1 grade (t=1.0)", "mech_consistency_t10"),
            ("P2 grade (t=1.0)", "nbhd_mech_consistency_t10"),
            ("P1 tier", "mavis_tier"), ("P2 tier", "nbhd_tier"),
            ("pipeline_agreement", "pipeline_agreement")]
    grade_headers = {"P1 grade (t=1.0)", "P2 grade (t=1.0)"}
    for i, (_, r) in enumerate(diff.iterrows()):
        for j, (h, c) in enumerate(cols, start=1):
            v = r.get(c, "")
            if pd.isna(v): v = ""
            cell = ws.cell(row=6 + i, column=j, value=v)
            style_body(cell)
            if h in grade_headers and v:
                fill = fill_for_grade(v)
                if fill: cell.fill = fill
    ws.freeze_panes = "C6"
    auto_width(ws)


def build_concordance_view(wb, df, mode_A):
    ws = wb.create_sheet("Concordance_View")
    ws.cell(row=1, column=1, value="4-Axis Concordance + Signal/Consensus Split").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    note = ("4-axis concordance combines structural signal (tier + DDG) with "
            "external pathogenicity predictors (AlphaMissense + Franklin). The "
            "signal/consensus split decomposes this into pipeline-internal "
            "structural detection vs pipeline-external pathogenicity prediction. "
            "Concordance = signal + consensus by construction.")
    ws.cell(row=3, column=1, value=note).font = FONT_NOTE
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)

    headers = ["gene", "variant", "role" if mode_A else "system",
               "concordance_strict_full", "concordance_relaxed_full",
               "structural_signal_strict (P1)", "structural_signal_strict (P2)",
               "external_consensus_strict",
               "AM class", "Franklin"]
    write_header_row(ws, 5, headers)

    cols = [("gene", "gene"), ("variant", "variant"),
            ("role" if mode_A else "system", "role" if mode_A else "system"),
            ("concordance_strict_full", "concordance_strict_full"),
            ("concordance_relaxed_full", "concordance_relaxed_full"),
            ("structural_signal_strict (P1)", "structural_signal_strict"),
            ("structural_signal_strict (P2)", "nbhd_structural_signal_strict"),
            ("external_consensus_strict", "external_consensus_strict"),
            ("AM class", "AM class"), ("Franklin", "franklin")]
    for i, (_, r) in enumerate(df.iterrows()):
        for j, (h, c) in enumerate(cols, start=1):
            v = r.get(c, "")
            if pd.isna(v): v = ""
            cell = ws.cell(row=6 + i, column=j, value=v)
            style_body(cell)
    ws.freeze_panes = "C6"
    auto_width(ws)


def build_methods(wb, mode_A):
    ws = wb.create_sheet("Methods_Definitions")
    ws.cell(row=1, column=1, value="Methods and Column Definitions").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    blocks = [
        ("What this report contains",
         "This report summarizes per-variant mechanism consistency grading from "
         "the MAVIS v7 structural variant pipeline. The pipeline detects "
         "structural disruption from missense variants; pathogenicity is "
         "evaluated separately. See MAVIS_v7_grading_rubric_v2.md for full "
         "rubric documentation."),
        ("Mechanism class derivation",
         "Each variant is classified into one of: structurally_silent, "
         "fold_mechanism, ppi_destab_mechanism, ppi_stab_mechanism, "
         "mixed_structural, interface_uncommitted_magnitude, "
         "structurally_uncommitted. Derivation is role-independent (does NOT use "
         "pathogenic/benign labels) — based on the four expected_* axis annotations "
         "from literature curation, with evidence-aware mild_destab handling."),
        ("Threshold methodology",
         "Mechanism consistency is computed at four ΔΔG thresholds: t=1.0, 1.5, 2.0, "
         "2.5 kcal/mol. These reflect the field's range of accepted thresholds: "
         "1.0 (CAGI5 frataxin convention; prior CHD work continuity), "
         "1.5 (Caldararu/Guerois empirical classification optimum), "
         "2.0 (intermediate), "
         "2.5 (Sapozhnikov 2023 confident-detection threshold given ±2.9 kcal/mol "
         "FoldX-vs-experimental uncertainty bound). No single threshold is "
         "designated 'primary'; the t=2.5 grade is the headline summary."),
        ("Grade definitions",
         "consistent: all positive axes correctly called, no false-fires on negative axes. "
         "partial: primary positive axis correct but secondary issue (FP on negative axis, "
         "or magnitude miss on positive axis). "
         "inconsistent: missed primary positive axis, direction flipped, or any "
         "false-firing on a structurally_silent variant. "
         "N/A: structure unevaluable, all axes uncommitted, or insufficient ground truth."),
        ("CI methodology",
         "Internal CI: ±1.96 × FoldX run-to-run SD across 5 BuildModel/AnalyseComplex "
         "replicates. Captures run-to-run reproducibility. Tight, optimistic. "
         "Sapozhnikov CI: ±2.9 kcal/mol (fold) or ±3.5 kcal/mol (binding) per "
         "Sapozhnikov et al. 2023 BMC Bioinformatics global 95% prediction interval. "
         "Captures structure quality + conformational + biochemical uncertainty. "
         "Wide, realistic. The two CI flavors answer different questions and are "
         "reported in parallel."),
        ("Threshold-stable flag",
         "True if a variant's mechanism consistency grade is identical across all "
         "four threshold values (t=1.0, 1.5, 2.0, 2.5 kcal/mol). Variants flagged "
         "False have grades that depend on threshold choice — these warrant "
         "individual examination in the Discussion."),
        ("Pipeline 1 vs Pipeline 2",
         "Pipeline 1 (single-residue): mechanism call uses only the variant residue's "
         "ΔΔG and tier. Pipeline 2 (neighborhood): mechanism call considers a "
         "spatial window around the variant residue, capturing regional structural "
         "context. The two often agree but differ on ~26 of 44 variants in "
         "specific calls; mechanism consistency scores are similar but the per-variant "
         "breakdown differs."),
        ("Mode A vs Mode B",
         "Mode A: full benchmark mode — ground-truth annotations present, all sheets "
         "populated. Mode B: production mode — no ground truth, mechanism consistency "
         "grading is skipped, structural signal and CIs still emitted."),
        ("Key citations",
         "Sapozhnikov et al. 2023 BMC Bioinformatics 24:426 — FoldX uncertainty quantification. "
         "Caldararu et al. 2020 Sci Rep 10:14245 — pathogenicity classification optimum. "
         "Guerois et al. 2002 J Mol Biol 320:369 — original FoldX parameterization. "
         "Hauser et al. 2021 BMC Bioinformatics 22:107 — structural sensitivity of stability predictors."),
    ]
    if not mode_A:
        blocks.insert(1, ("Operating mode",
                           "This report was generated in Mode B (no ground truth). "
                           "Mechanism consistency grading sheets are blank or omitted; "
                           "structural signal and CI sheets are fully populated."))

    row = 3
    for title, body in blocks:
        c = ws.cell(row=row, column=1, value=title)
        c.font = FONT_BODY_B
        c.fill = FILL_SECTION
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c2 = ws.cell(row=row + 1, column=1, value=body)
        c2.font = FONT_BODY
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=4)
        ws.row_dimensions[row + 1].height = 60
        row += 3

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30


def build_caveats(wb, df, mode_A):
    ws = wb.create_sheet("Notes_and_Caveats")
    ws.cell(row=1, column=1, value="Notes, Caveats, Verification Status").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    blocks = [
        ("Verification status",
         "Per-variant ground-truth annotations are based on literature curation "
         "from the source benchmark. A paper-batched verification pass against "
         "primary sources is in progress. Variants with explicit annotations in "
         "the 'notes' column on the Per_Variant sheet have been documented; "
         "verification of specific quantitative claims may flag updates that "
         "require pipeline re-runs."),
        ("Known caveats — mild_destab annotation",
         "Five variants (TNNI3 R145G/Q, CALM1 D96V/N98S/F142L) carry 'mild_destab' "
         "expected ΔΔG annotations. The pipeline applies evidence-aware tokenization: "
         "mild_destab counts as positive only when evidence_axes is structural; "
         "mild_destab with functional/population evidence is treated as not-tested."),
        ("Known caveats — structurally_silent specificity",
         "structurally_silent is the largest mech class (24/44 variants). Of these, "
         "11 grade consistent at t=1.0 and 13 grade inconsistent — pipeline over-fires "
         "on roughly half. This is the major specificity finding and is consistent "
         "with FoldX's known uncertainty band per Sapozhnikov 2023."),
        ("Mechanism control notes",
         "11 variants are tagged 'mechanism_control' in the source benchmark. Under "
         "the v2 grading rubric, role does not enter mechanism consistency grading; "
         "these variants are graded against their structural annotations like any "
         "other. The label is informational only."),
        ("Pipeline 1 vs 2 trade-off",
         "P2 (neighborhood) shows higher sensitivity and slightly lower specificity "
         "than P1 (single-residue). Mechanism consistency scores at t=2.5 are "
         "identical (0.773), but the underlying calls differ on 26 variants. "
         "Choice between pipelines depends on use case (sensitivity-favoring vs "
         "specificity-favoring)."),
        ("Threshold choice rationale",
         "We deliberately do not designate a single 'primary' threshold. The "
         "field has not converged: ROC-optimization studies cluster at 1.5–1.6 "
         "kcal/mol; uncertainty-quantification studies push toward 2.5+ kcal/mol; "
         "many recent papers use 1.0 or 1.6 per their specific benchmark needs. "
         "The four-threshold sweep IS the principal output."),
    ]
    if mode_A:
        n_unstable = int((~df["mech_consistency_threshold_stable"]).sum())
        blocks.append(("Threshold-sensitive variants (this dataset)",
                       f"{n_unstable} of 44 variants (P1) have grades that change "
                       f"across the four thresholds. These appear on the "
                       f"Threshold_Sensitivity sheet and warrant individual "
                       f"examination during paper writing."))

    row = 3
    for title, body in blocks:
        c = ws.cell(row=row, column=1, value=title)
        c.font = FONT_BODY_B
        c.fill = FILL_SECTION
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c2 = ws.cell(row=row + 1, column=1, value=body)
        c2.font = FONT_BODY
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=4)
        ws.row_dimensions[row + 1].height = 60
        row += 3

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30


# ============================================================================
# Main
# ============================================================================
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--concordance", default="results/mavis_v7_concordance.csv",
                    help="Path to concordance CSV from apply_concordance_v4.py")
    ap.add_argument("--out", default="results/mavis_v7_report.xlsx",
                    help="Output XLSX path")
    args = ap.parse_args()

    cp = Path(args.concordance)
    op = Path(args.out)
    if not cp.exists():
        print(f"ERROR: concordance CSV not found: {cp}", file=sys.stderr)
        sys.exit(1)
    op.parent.mkdir(parents=True, exist_ok=True)

    print(f"Reading {cp}...")
    df = pd.read_csv(cp)
    print(f"  {len(df)} variants, {len(df.columns)} columns")

    # Detect mode
    mode_A = ("expected_mech_class" in df.columns
              and "role" in df.columns
              and "mech_consistency_t10" in df.columns)
    print(f"  Mode: {'A (benchmark with ground truth)' if mode_A else 'B (production / structural-signal only)'}")

    # Build workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("\nBuilding sheets:")
    print("  1/9: Per_Variant")
    build_per_variant(wb, df, mode_A)
    print("  2/9: Mech_Consistency_Summary")
    build_mech_consistency_summary(wb, df, mode_A)
    print("  3/9: Threshold_Sensitivity")
    build_threshold_sensitivity(wb, df, mode_A)
    print("  4/9: DDG_Distinguishability")
    build_distinguishability(wb, df)
    print("  5/9: Signal_in_Benign")
    build_signal_in_benign(wb, df, mode_A)
    print("  6/9: P1_vs_P2_Comparison")
    build_p1_vs_p2(wb, df, mode_A)
    print("  7/9: Concordance_View")
    build_concordance_view(wb, df, mode_A)
    print("  8/9: Methods_Definitions")
    build_methods(wb, mode_A)
    print("  9/9: Notes_and_Caveats")
    build_caveats(wb, df, mode_A)

    wb.save(op)
    print(f"\nWrote {op}")


if __name__ == "__main__":
    main()
